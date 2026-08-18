#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
工厂项目管理系统 - Flask 应用主文件
管理订单从打样到出货的全流程 + 物料库存管理 + 产能预期 + 打卡功能
"""

import os
import sys
import shutil
import io
from datetime import datetime, date, timedelta
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, session, g
from functools import wraps

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from models import db, User, Order, ProcessStep, Material, InventoryRecord, OrderMaterial, CapacityPlan, StepDefinition, CheckinRecord, DEFAULT_STEPS

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'local-only-change-me')
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('COOKIE_SECURE', '').lower() == 'true'

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, 'factory.db')
BACKUP_DIR = os.path.join(BASE_DIR, 'backups')
# Keep SQLite for local use, but allow a hosted PostgreSQL database for
# multi-user deployments. SQLite files are not suitable for concurrent writes
# on a web host with more than one worker.
database_url = os.environ.get('DATABASE_URL', '').strip()
if database_url.startswith('postgres://'):
    database_url = database_url.replace('postgres://', 'postgresql+psycopg://', 1)
elif database_url.startswith('postgresql://'):
    database_url = database_url.replace('postgresql://', 'postgresql+psycopg://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = database_url or f'sqlite:///{DB_PATH}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# 确保备份目录存在
os.makedirs(BACKUP_DIR, exist_ok=True)

db.init_app(app)


@app.before_request
def load_current_user_and_protect_writes():
    """Anonymous visitors can monitor production; writes require a login."""
    g.user = None
    user_id = session.get('user_id')
    if user_id:
        g.user = db.session.get(User, user_id)
        if not g.user or not g.user.is_active:
            session.clear()
            g.user = None

    if request.method in ('POST', 'PUT', 'PATCH', 'DELETE'):
        if request.endpoint != 'login' and g.user is None:
            if request.is_json or request.path.startswith('/api/'):
                return jsonify({'success': False, 'error': '请先登录后再修改数据'}), 401
            flash('查看无需登录，修改数据请先登录', 'warning')
            return redirect(url_for('login', next=request.path))


def admin_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not g.user or g.user.role != 'admin':
            flash('需要管理员权限', 'danger')
            return redirect(url_for('index'))
        return view(*args, **kwargs)
    return wrapped


@app.context_processor
def inject_current_user():
    return {'current_user': g.get('user')}


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        user = User.query.filter_by(username=username, is_active=True).first()
        if user and user.check_password(password):
            session.clear()
            session['user_id'] = user.id
            next_url = request.args.get('next') or request.form.get('next') or url_for('index')
            return redirect(next_url if next_url.startswith('/') else url_for('index'))
        flash('账号或密码不正确', 'danger')
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    flash('已安全退出', 'success')
    return redirect(url_for('index'))


@app.route('/users', methods=['GET'])
@admin_required
def users_page():
    return render_template('users.html', users=User.query.order_by(User.created_at.desc()).all())


@app.route('/api/users/add', methods=['POST'])
@admin_required
def user_add():
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '')
    display_name = request.form.get('display_name', '').strip()
    role = request.form.get('role', 'editor')
    if not username or len(password) < 8:
        flash('账号不能为空，密码至少需要 8 位', 'danger')
        return redirect(url_for('users_page'))
    if User.query.filter_by(username=username).first():
        flash('该账号已存在', 'danger')
        return redirect(url_for('users_page'))
    user = User(username=username, display_name=display_name, role='admin' if role == 'admin' else 'editor')
    user.set_password(password)
    db.session.add(user)
    db.session.commit()
    flash(f'账号 {username} 创建成功', 'success')
    return redirect(url_for('users_page'))


# ==================== 工具函数 ====================

def get_status_class(status):
    mapping = {
        'pending': 'secondary',
        'in_progress': 'primary',
        'completed': 'success',
        'skipped': 'warning',
        'draft': 'secondary',
    }
    return mapping.get(status, 'secondary')


def get_status_label(status):
    mapping = {
        'pending': '待处理',
        'in_progress': '进行中',
        'completed': '已完成',
        'skipped': '已跳过',
        'draft': '草稿',
    }
    return mapping.get(status, status)


def get_category_label(cat):
    mapping = {
        'sampling': '打样阶段',
        'common': '准备阶段',
        'production': '生产阶段',
        'inspection': '验货',
        'testing': '测试',
    }
    return mapping.get(cat, cat)


# ==================== 路由 - 首页/订单列表 ====================

@app.route('/')
def index():
    status_filter = request.args.get('status', '')
    search = request.args.get('search', '')

    query = Order.query.filter_by(is_active=True)

    if search:
        query = query.filter(
            db.or_(
                Order.order_no.contains(search),
                Order.customer_name.contains(search),
                Order.product_name.contains(search),
            )
        )

    orders = query.order_by(Order.created_at.desc()).all()

    if status_filter:
        orders = [o for o in orders if o.current_status() == status_filter]

    return render_template('index.html',
                           orders=orders,
                           status_filter=status_filter,
                           search=search,
                           get_status_class=get_status_class,
                           get_status_label=get_status_label,
                           get_category_label=get_category_label)


# ==================== 路由 - 创建订单 ====================

@app.route('/order/create', methods=['GET', 'POST'])
def order_create():
    if request.method == 'POST':
        try:
            order_no = request.form.get('order_no', '').strip()
            if not order_no:
                flash('请输入订单号', 'danger')
                return render_template('order_form.html', order=None)

            exist = Order.query.filter_by(order_no=order_no).first()
            if exist:
                flash(f'订单号 "{order_no}" 已存在', 'danger')
                return render_template('order_form.html', order=None)

            order = Order(
                order_no=order_no,
                customer_name=request.form.get('customer_name', '').strip(),
                product_name=request.form.get('product_name', '').strip(),
                order_date=parse_date(request.form.get('order_date')),
                delivery_date=parse_date(request.form.get('delivery_date')),
                quantity=int(request.form.get('quantity', 0) or 0),
                notes=request.form.get('notes', '').strip(),
            )
            db.session.add(order)
            db.session.flush()

            for step_def in DEFAULT_STEPS:
                step = ProcessStep(
                    order_id=order.id,
                    seq=step_def['seq'],
                    category=step_def['category'],
                    step_name=step_def['step_name'],
                    step_type='' if step_def.get('has_type') else None,
                )
                db.session.add(step)

            db.session.commit()
            flash(f'订单 "{order_no}" 创建成功！', 'success')
            return redirect(url_for('order_detail', order_id=order.id))

        except Exception as e:
            db.session.rollback()
            flash(f'创建失败: {str(e)}', 'danger')

    return render_template('order_form.html', order=None)


# ==================== 路由 - 订单详情 ====================

@app.route('/order/<int:order_id>')
def order_detail(order_id):
    order = Order.query.get_or_404(order_id)
    steps_by_category = {}
    for s in order.steps:
        steps_by_category.setdefault(s.category, []).append(s)

    materials = Material.query.filter_by(is_active=True).order_by(Material.name).all()

    return render_template('order_detail.html',
                           order=order,
                           steps_by_category=steps_by_category,
                           materials=materials,
                           get_status_class=get_status_class,
                           get_status_label=get_status_label,
                           get_category_label=get_category_label)


# ==================== 路由 - 编辑订单 ====================

@app.route('/order/<int:order_id>/edit', methods=['GET', 'POST'])
def order_edit(order_id):
    order = Order.query.get_or_404(order_id)

    if request.method == 'POST':
        try:
            order.customer_name = request.form.get('customer_name', '').strip()
            order.product_name = request.form.get('product_name', '').strip()
            order.order_date = parse_date(request.form.get('order_date'))
            order.delivery_date = parse_date(request.form.get('delivery_date'))
            order.quantity = int(request.form.get('quantity', 0) or 0)
            order.notes = request.form.get('notes', '').strip()
            db.session.commit()
            flash('订单信息已更新', 'success')
            return redirect(url_for('order_detail', order_id=order.id))
        except Exception as e:
            db.session.rollback()
            flash(f'更新失败: {str(e)}', 'danger')

    return render_template('order_form.html', order=order)


# ==================== 路由 - 更新步骤状态 ====================

@app.route('/api/step/<int:step_id>/update', methods=['POST'])
def step_update(step_id):
    step = ProcessStep.query.get_or_404(step_id)
    try:
        data = request.json or request.form

        new_status = data.get('status', step.status)
        step.status = new_status

        if new_status == 'completed' and not step.completed_date:
            step.completed_date = date.today()
        elif new_status in ('pending', 'in_progress'):
            step.completed_date = None

        if 'step_type' in data:
            step.step_type = data['step_type']
        if 'assignee' in data:
            step.assignee = data['assignee']
        if 'notes' in data:
            step.notes = data['notes']
        if 'planned_date' in data and data['planned_date']:
            step.planned_date = parse_date(data['planned_date'])

        db.session.commit()

        order = Order.query.get(step.order_id)
        if order and order.current_status() == 'completed':
            flash(f'订单 "{order.order_no}" 所有步骤已完成！', 'success')

        if request.is_json:
            return jsonify({'success': True, 'status': new_status})

        flash('步骤已更新', 'success')
        return redirect(url_for('order_detail', order_id=step.order_id))

    except Exception as e:
        db.session.rollback()
        msg = f'更新失败: {str(e)}'
        if request.is_json:
            return jsonify({'success': False, 'error': msg}), 400
        flash(msg, 'danger')
        return redirect(url_for('order_detail', order_id=step.order_id))


# ==================== 路由 - 删除订单 ====================

@app.route('/order/<int:order_id>/delete', methods=['POST'])
def order_delete(order_id):
    order = Order.query.get_or_404(order_id)
    try:
        order.is_active = False
        db.session.commit()
        flash(f'订单 "{order.order_no}" 已删除', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'删除失败: {str(e)}', 'danger')
    return redirect(url_for('index'))


# ==================== 路由 - 添加自定义步骤 ====================

@app.route('/order/<int:order_id>/step/add', methods=['POST'])
def order_add_step(order_id):
    order = Order.query.get_or_404(order_id)
    try:
        step_name = request.form.get('step_name', '').strip()
        category = request.form.get('category', 'common')
        if not step_name:
            flash('请输入步骤名称', 'danger')
            return redirect(url_for('order_detail', order_id=order.id))

        max_seq = max((s.seq for s in order.steps), default=0)
        step = ProcessStep(
            order_id=order.id,
            seq=max_seq + 1,
            category=category,
            step_name=step_name,
        )
        db.session.add(step)
        db.session.commit()
        flash(f'已添加步骤: {step_name}', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'添加失败: {str(e)}', 'danger')
    return redirect(url_for('order_detail', order_id=order.id))


# ============================================================
#  🧱 物料管理路由
# ============================================================

@app.route('/materials')
def material_list():
    category = request.args.get('category', '')
    search = request.args.get('search', '')

    query = Material.query.filter_by(is_active=True)

    if category:
        query = query.filter_by(category=category)
    if search:
        query = query.filter(Material.name.contains(search))

    materials = query.order_by(Material.category, Material.name).all()
    categories = db.session.query(Material.category).filter_by(is_active=True).distinct().all()
    categories = [c[0] for c in categories if c[0]]

    return render_template('materials.html',
                           materials=materials,
                           categories=categories,
                           current_category=category,
                           search=search)


@app.route('/api/material/add', methods=['POST'])
def material_add():
    try:
        data = request.json
        name = data.get('name', '').strip()
        if not name:
            return jsonify({'success': False, 'error': '请输入物料名称'}), 400

        material = Material(
            name=name,
            spec=data.get('spec', '').strip(),
            unit=data.get('unit', '个'),
            category=data.get('category', '包材'),
            stock_qty=float(data.get('stock_qty', 0) or 0),
            min_stock=float(data.get('min_stock', 0) or 0),
            note=data.get('note', ''),
        )
        db.session.add(material)
        db.session.commit()

        if material.stock_qty > 0:
            record = InventoryRecord(
                material_id=material.id,
                type='purchase',
                quantity=material.stock_qty,
                before_qty=0,
                after_qty=material.stock_qty,
                note='初始化库存',
            )
            db.session.add(record)
            db.session.commit()

        return jsonify({'success': True, 'material': material.to_dict()})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/material/<int:material_id>/update', methods=['POST'])
def material_update(material_id):
    material = Material.query.get_or_404(material_id)
    try:
        data = request.json
        if 'name' in data: material.name = data['name']
        if 'spec' in data: material.spec = data['spec']
        if 'unit' in data: material.unit = data['unit']
        if 'category' in data: material.category = data['category']
        if 'min_stock' in data: material.min_stock = float(data['min_stock'] or 0)
        if 'note' in data: material.note = data['note']

        db.session.commit()
        return jsonify({'success': True, 'material': material.to_dict()})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/material/<int:material_id>/delete', methods=['POST'])
def material_delete(material_id):
    material = Material.query.get_or_404(material_id)
    try:
        material.is_active = False
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/material/<int:material_id>/stock', methods=['POST'])
def material_stock_change(material_id):
    material = Material.query.get_or_404(material_id)
    try:
        data = request.json
        op_type = data.get('type', '')
        qty = float(data.get('quantity', 0))
        if qty <= 0:
            return jsonify({'success': False, 'error': '数量必须大于0'}), 400

        before = material.stock_qty
        if op_type == 'purchase':
            material.stock_qty += qty
            type_label = '购进'
        elif op_type == 'consume':
            if material.stock_qty < qty:
                return jsonify({'success': False, 'error': f'库存不足！当前库存: {material.stock_qty} {material.unit}'}), 400
            material.stock_qty -= qty
            type_label = '消耗'
        else:
            return jsonify({'success': False, 'error': '无效的操作类型'}), 400

        after = material.stock_qty

        record = InventoryRecord(
            material_id=material.id,
            type=op_type,
            quantity=qty,
            before_qty=before,
            after_qty=after,
            operator=data.get('operator', ''),
            note=data.get('note', f'{type_label} {qty}{material.unit}'),
        )
        db.session.add(record)
        db.session.commit()

        return jsonify({'success': True, 'material': material.to_dict(), 'record': record.to_dict()})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400


# ============================================================
#  📄 库存流水路由
# ============================================================

@app.route('/inventory/records')
def inventory_records():
    material_id = request.args.get('material_id', '')
    query = InventoryRecord.query.order_by(InventoryRecord.created_at.desc())

    if material_id:
        query = query.filter_by(material_id=int(material_id))

    records = query.limit(200).all()
    materials = Material.query.filter_by(is_active=True).order_by(Material.name).all()

    for r in records:
        mat = Material.query.get(r.material_id)
        r.material_name = mat.name if mat else '已删除'
        r.material_unit = mat.unit if mat else ''

    return render_template('inventory_records.html',
                           records=records,
                           materials=materials,
                           selected_material_id=int(material_id) if material_id else '')


# ============================================================
#  📎 订单物料关联
# ============================================================

@app.route('/api/order/<int:order_id>/material/add', methods=['POST'])
def order_material_add(order_id):
    order = Order.query.get_or_404(order_id)
    try:
        data = request.json
        material_id = data.get('material_id')
        required_qty = float(data.get('required_qty', 0) or 0)

        if not material_id:
            return jsonify({'success': False, 'error': '请选择物料'}), 400

        exist = OrderMaterial.query.filter_by(order_id=order.id, material_id=material_id).first()
        if exist:
            exist.required_qty += required_qty
        else:
            om = OrderMaterial(
                order_id=order.id,
                material_id=material_id,
                required_qty=required_qty,
            )
            db.session.add(om)

        db.session.commit()
        return jsonify({'success': True})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/order/<int:order_id>/materials')
def order_materials_list(order_id):
    order = Order.query.get_or_404(order_id)
    return jsonify([om.to_dict() for om in order.materials])


@app.route('/api/order/material/<int:om_id>/delete', methods=['POST'])
def order_material_delete(om_id):
    om = OrderMaterial.query.get_or_404(om_id)
    try:
        db.session.delete(om)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400


# ============================================================
#  💾 数据库备份
# ============================================================

@app.route('/api/backup', methods=['POST'])
def backup_database():
    """一键备份数据库"""
    try:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_filename = f'factory_backup_{timestamp}.db'
        backup_path = os.path.join(BACKUP_DIR, backup_filename)

        if os.path.exists(DB_PATH):
            shutil.copy2(DB_PATH, backup_path)
            return jsonify({
                'success': True,
                'filename': backup_filename,
                'path': backup_path,
                'message': f'备份成功！文件: {backup_filename}'
            })
        else:
            return jsonify({'success': False, 'error': '数据库文件不存在'}), 400

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/backup/list')
def backup_list():
    """查看备份列表"""
    try:
        backups = []
        if os.path.exists(BACKUP_DIR):
            for f in sorted(os.listdir(BACKUP_DIR), reverse=True):
                if f.endswith('.db'):
                    fpath = os.path.join(BACKUP_DIR, f)
                    size = os.path.getsize(fpath)
                    backups.append({
                        'filename': f,
                        'size': f'{size / 1024:.1f} KB',
                        'time': datetime.fromtimestamp(os.path.getmtime(fpath)).strftime('%Y-%m-%d %H:%M:%S')
                    })
        return jsonify({'success': True, 'backups': backups})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


# ============================================================
#  📊 导出 Excel
# ============================================================

def export_orders_to_excel():
    """导出订单列表到 Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()

    # ---- Sheet 1: 订单列表 ----
    ws = wb.active
    ws.title = '订单列表'
    headers = ['订单号', '客户名称', '产品名称', '数量', '下单日期', '交货日期', '当前状态', '备注']
    ws.append(headers)

    # 表头样式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    orders = Order.query.filter_by(is_active=True).order_by(Order.created_at.desc()).all()
    for o in orders:
        ws.append([
            o.order_no, o.customer_name, o.product_name, o.quantity,
            str(o.order_date or ''), str(o.delivery_date or ''),
            get_status_label(o.current_status()), o.notes or ''
        ])

    # 自动调整列宽
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 4, 12)

    # ---- Sheet 2: 物料库存 ----
    ws2 = wb.create_sheet('物料库存')
    ws2.append(['物料名称', '规格', '类别', '单位', '当前库存', '最低警戒', '备注'])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font

    materials = Material.query.filter_by(is_active=True).order_by(Material.category, Material.name).all()
    for m in materials:
        ws2.append([m.name, m.spec or '', m.category, m.unit, m.stock_qty, m.min_stock or 0, m.note or ''])

    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws2.column_dimensions[col[0].column_letter].width = max(max_len + 4, 12)

    # ---- Sheet 3: 库存流水 ----
    ws3 = wb.create_sheet('库存流水')
    ws3.append(['时间', '物料', '操作类型', '数量', '操作前库存', '操作后库存', '备注'])
    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = header_font

    records = InventoryRecord.query.order_by(InventoryRecord.created_at.desc()).limit(500).all()
    for r in records:
        mat = Material.query.get(r.material_id)
        mat_name = mat.name if mat else '已删除'
        type_label = '购进' if r.type == 'purchase' else '消耗'
        ws3.append([
            r.created_at.strftime('%Y-%m-%d %H:%M') if r.created_at else '',
            mat_name, type_label, r.quantity, r.before_qty, r.after_qty, r.note or ''
        ])

    for col in ws3.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws3.column_dimensions[col[0].column_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@app.route('/export/orders')
def export_orders():
    """导出 Excel 文件"""
    try:
        output = export_orders_to_excel()
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'工厂数据_{timestamp}.xlsx'
        )
    except Exception as e:
        flash(f'导出失败: {str(e)}', 'danger')
        return redirect(request.referrer or url_for('index'))


# ============================================================
#  ✅ 打卡功能
# ============================================================

@app.route('/checkin')
def checkin_page():
    """打卡主页"""
    orders = Order.query.filter_by(is_active=True).order_by(Order.created_at.desc()).all()
    step_defs = StepDefinition.query.filter_by(is_active=True).order_by(StepDefinition.category, StepDefinition.name).all()
    records = CheckinRecord.query.order_by(CheckinRecord.created_at.desc()).all()
    return render_template('checkin.html', orders=orders, step_defs=step_defs, records=records)


# --- 步骤定义管理 ---

@app.route('/api/step-definitions', methods=['GET'])
def api_step_definitions():
    """获取所有步骤定义"""
    defs = StepDefinition.query.filter_by(is_active=True).order_by(StepDefinition.category, StepDefinition.name).all()
    return jsonify([d.to_dict() for d in defs])


@app.route('/api/step-definitions/add', methods=['POST'])
def api_step_definition_add():
    """添加步骤定义"""
    try:
        data = request.json
        name = data.get('name', '').strip()
        category = data.get('category', 'production').strip()
        if not name:
            return jsonify({'success': False, 'error': '请输入步骤名称'}), 400

        exist = StepDefinition.query.filter_by(name=name, is_active=True).first()
        if exist:
            return jsonify({'success': False, 'error': f'步骤 "{name}" 已存在'}), 400

        sd = StepDefinition(name=name, category=category)
        db.session.add(sd)
        db.session.commit()
        return jsonify({'success': True, 'step': sd.to_dict()})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/step-definitions/<int:sd_id>/delete', methods=['POST'])
def api_step_definition_delete(sd_id):
    """删除步骤定义"""
    sd = StepDefinition.query.get_or_404(sd_id)
    try:
        sd.is_active = False
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400


# --- 打卡操作 ---

@app.route('/api/checkin/save', methods=['POST'])
def api_checkin_save():
    """保存打卡记录，同时同步产能预期"""
    try:
        data = request.json
        order_id = data.get('order_id')
        step_name = data.get('step_name', '').strip()
        status = data.get('status', 'in_progress')
        completed_qty = float(data.get('completed_qty', 0) or 0)
        operator = data.get('operator', '').strip()

        if not order_id:
            return jsonify({'success': False, 'error': '请选择订单'}), 400
        if not step_name:
            return jsonify({'success': False, 'error': '请选择步骤'}), 400

        order = Order.query.get(order_id)
        if not order:
            return jsonify({'success': False, 'error': '订单不存在'}), 404

        # 创建打卡记录
        record = CheckinRecord(
            order_id=order_id,
            step_name=step_name,
            status=status,
            completed_qty=completed_qty,
            operator=operator,
            checkin_date=date.today(),
        )
        db.session.add(record)

        # 同步到产能预期：找到对应订单和产品名称的产能记录，更新已完成数量
        if completed_qty > 0:
            # 查找匹配的产能记录（按 order_id + 步骤名称）
            plan = CapacityPlan.query.filter_by(
                order_id=order_id,
                note=step_name  # 产能记录的 note 字段存步骤名称
            ).first()
            if not plan:
                # 尝试找到该订单的任意产能记录
                plan = CapacityPlan.query.filter_by(order_id=order_id).first()

            if plan:
                plan.completed_qty = (plan.completed_qty or 0) + completed_qty

        db.session.commit()

        return jsonify({'success': True, 'record': record.to_dict()})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/checkin/records')
def api_checkin_records():
    """获取打卡记录（可按订单筛选）"""
    order_id = request.args.get('order_id', '')
    query = CheckinRecord.query.order_by(CheckinRecord.created_at.desc())

    if order_id:
        query = query.filter_by(order_id=int(order_id))

    records = query.limit(200).all()
    return jsonify([r.to_dict() for r in records])


@app.route('/api/checkin/records/<int:record_id>/delete', methods=['POST'])
def api_checkin_record_delete(record_id):
    """删除打卡记录"""
    record = CheckinRecord.query.get_or_404(record_id)
    try:
        # 回退产能预期的已完成数量
        if record.completed_qty > 0:
            plan = CapacityPlan.query.filter_by(
                order_id=record.order_id,
                note=record.step_name
            ).first()
            if plan:
                plan.completed_qty = max(0, (plan.completed_qty or 0) - record.completed_qty)

        db.session.delete(record)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400


# ============================================================
#  🏭 产能预期
# ============================================================

@app.route('/capacity')
def capacity_plan():
    """产能预期页面"""
    order_id = request.args.get('order_id', '')
    orders = Order.query.filter_by(is_active=True).order_by(Order.created_at.desc()).all()
    query = CapacityPlan.query.order_by(CapacityPlan.created_at.desc())
    if order_id:
        query = query.filter_by(order_id=int(order_id))
    plans = query.all()
    return render_template('capacity.html', orders=orders, plans=plans, selected_order_id=order_id)


@app.route('/api/capacity/calculate', methods=['POST'])
def capacity_calculate():
    """计算产能（AJAX，不保存）"""
    try:
        data = request.json
        total_qty = float(data.get('total_qty', 0))
        hourly_rate = float(data.get('hourly_rate', 0))
        workers = float(data.get('workers', 1))
        hours_per_day = float(data.get('hours_per_day', 8))
        start_date_str = data.get('start_date', '')

        if total_qty <= 0 or hourly_rate <= 0:
            return jsonify({'success': False, 'error': '总数量和小时产能必须大于0'}), 400

        daily_output = hourly_rate * workers * hours_per_day
        est_days = total_qty / daily_output if daily_output > 0 else 0

        est_end_date = None
        if start_date_str:
            from dateutil.parser import parse as dt_parse
            try:
                sd = dt_parse(start_date_str).date()
                est_end_date = sd + timedelta(days=int(est_days) + (1 if est_days > int(est_days) else 0))
            except:
                pass

        return jsonify({
            'success': True,
            'daily_output': round(daily_output, 1),
            'est_days': round(est_days, 1),
            'est_end_date': est_end_date.strftime('%Y-%m-%d') if est_end_date else '',
            'est_days_int': int(est_days) + (1 if est_days > int(est_days) else 0),
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/capacity/save', methods=['POST'])
def capacity_save():
    """保存产能预期记录"""
    try:
        data = request.json
        order_id = data.get('order_id') or None
        start_date = parse_date(data.get('start_date'))
        hourly_rate = float(data.get('hourly_rate', 0))
        workers = float(data.get('workers', 1))
        hours_per_day = float(data.get('hours_per_day', 8))
        total_qty = float(data.get('total_qty', 0))

        daily_output = hourly_rate * workers * hours_per_day
        est_days = total_qty / daily_output if daily_output > 0 else 0

        end_date = None
        if start_date and est_days > 0:
            end_date = start_date + timedelta(days=int(est_days) + (1 if est_days > int(est_days) else 0))

        plan = CapacityPlan(
            order_id=order_id,
            product_name=data.get('product_name', ''),
            total_qty=total_qty,
            hourly_rate=hourly_rate,
            workers=workers,
            hours_per_day=hours_per_day,
            daily_output=round(daily_output, 1),
            est_days=round(est_days, 1),
            start_date=start_date,
            est_end_date=end_date,
            note=data.get('note', ''),
        )
        db.session.add(plan)
        db.session.commit()

        return jsonify({'success': True, 'plan': plan.to_dict()})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/capacity/delete/<int:plan_id>', methods=['POST'])
def capacity_delete(plan_id):
    """删除产能记录"""
    plan = CapacityPlan.query.get_or_404(plan_id)
    try:
        db.session.delete(plan)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/capacity/order/<int:order_id>')
def capacity_order_plans(order_id):
    """获取某订单的所有产能记录"""
    plans = CapacityPlan.query.filter_by(order_id=order_id).order_by(CapacityPlan.created_at).all()
    return jsonify([p.to_dict() for p in plans])


@app.route('/api/capacity/batch/calculate', methods=['POST'])
def capacity_batch_calculate():
    """计算多个工序的累计产能"""
    try:
        data = request.json
        segments = data.get('segments', [])
        if not segments:
            return jsonify({'success': False, 'error': '请至少添加一个工序'}), 400

        cumulative_days = 0
        overall_start_date = None
        segment_results = []

        for seg in segments:
            total_qty = float(seg.get('total_qty', 0))
            hourly_rate = float(seg.get('hourly_rate', 0))
            workers = float(seg.get('workers', 1))
            hours_per_day = float(seg.get('hours_per_day', 8))

            if total_qty <= 0 or hourly_rate <= 0:
                continue

            daily_output = hourly_rate * workers * hours_per_day
            est_days = total_qty / daily_output if daily_output > 0 else 0
            est_days_ceil = int(est_days) + (1 if est_days > int(est_days) else 0)
            cumulative_days += est_days_ceil

            segment_results.append({
                'process_name': seg.get('process_name', ''),
                'total_qty': total_qty,
                'hourly_rate': hourly_rate,
                'workers': workers,
                'hours_per_day': hours_per_day,
                'daily_output': round(daily_output, 1),
                'est_days': round(est_days, 1),
                'est_days_ceil': est_days_ceil,
            })

        if data.get('start_date'):
            from dateutil.parser import parse as dt_parse
            try:
                overall_start_date = dt_parse(data['start_date']).date()
            except:
                pass

        overall_end_date = None
        if overall_start_date and cumulative_days > 0:
            overall_end_date = overall_start_date + timedelta(days=cumulative_days)

        return jsonify({
            'success': True,
            'segments': segment_results,
            'cumulative_days': cumulative_days,
            'overall_start_date': overall_start_date.strftime('%Y-%m-%d') if overall_start_date else '',
            'overall_end_date': overall_end_date.strftime('%Y-%m-%d') if overall_end_date else '',
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


# ============================================================
#  📊 统计看板
# ============================================================

@app.route('/dashboard')
def dashboard():
    total = Order.query.filter_by(is_active=True).count()
    pending = len([o for o in Order.query.filter_by(is_active=True).all() if o.current_status() == 'pending'])
    in_progress = len([o for o in Order.query.filter_by(is_active=True).all() if o.current_status() == 'in_progress'])
    completed = len([o for o in Order.query.filter_by(is_active=True).all() if o.current_status() == 'completed'])

    recent_orders = Order.query.filter_by(is_active=True).order_by(Order.created_at.desc()).limit(5).all()

    materials = Material.query.filter_by(is_active=True).all()
    total_materials = len(materials)
    low_stock_materials = [m for m in materials if m.stock_qty <= m.min_stock and m.min_stock > 0]
    total_stock_value = sum(m.stock_qty for m in materials)

    category_stats = {}
    for m in materials:
        cat = m.category or '其他'
        if cat not in category_stats:
            category_stats[cat] = {'count': 0, 'total_stock': 0}
        category_stats[cat]['count'] += 1
        category_stats[cat]['total_stock'] += m.stock_qty

    recent_records = InventoryRecord.query.order_by(InventoryRecord.created_at.desc()).limit(10).all()
    for r in recent_records:
        mat = Material.query.get(r.material_id)
        r.material_name = mat.name if mat else '已删除'
        r.material_unit = mat.unit if mat else ''

    backup_count = 0
    if os.path.exists(BACKUP_DIR):
        backup_count = len([f for f in os.listdir(BACKUP_DIR) if f.endswith('.db')])

    return render_template('dashboard.html',
                           total=total, pending=pending,
                           in_progress=in_progress, completed=completed,
                           recent_orders=recent_orders,
                           total_materials=total_materials,
                           low_stock_materials=low_stock_materials,
                           total_stock_value=total_stock_value,
                           category_stats=category_stats,
                           recent_records=recent_records,
                           backup_count=backup_count,
                           get_status_class=get_status_class,
                           get_status_label=get_status_label)


# ==================== 辅助函数 ====================

def parse_date(date_str):
    if not date_str or not date_str.strip():
        return None
    from dateutil.parser import parse as dt_parse
    try:
        return dt_parse(date_str).date()
    except:
        return None


# ==================== 模板过滤器 ====================

@app.template_filter('status_class')
def status_class_filter(status):
    return get_status_class(status)


@app.template_filter('status_label')
def status_label_filter(status):
    return get_status_label(status)


@app.template_filter('category_label')
def category_label_filter(cat):
    return get_category_label(cat)


# ==================== 启动 ====================

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        admin_username = os.environ.get('ADMIN_USERNAME')
        admin_password = os.environ.get('ADMIN_PASSWORD')
        if admin_username and admin_password and not User.query.filter_by(username=admin_username).first():
            admin = User(username=admin_username, display_name='系统管理员', role='admin')
            admin.set_password(admin_password)
            db.session.add(admin)
            db.session.commit()
        print('✓ 数据库已初始化')
    app.run(host='0.0.0.0', port=5001, debug=True)
